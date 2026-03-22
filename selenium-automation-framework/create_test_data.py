"""
Run this Python script once to generate a sample TestData.xlsx
for the LinkedIn Automation framework.

Usage:
    pip install openpyxl
    python create_test_data.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "JobSearch"

# Header row
headers = ["jobTitle", "location", "experienceLevel", "keyword"]
header_fill = PatternFill(start_color="0A66C2", end_color="0A66C2", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[cell.column_letter].width = 22

# Data rows
test_data = [
    ["Java Developer",       "Bangalore",  "Mid-Senior level", "Spring Boot"],
    ["QA Automation Engineer","Remote",    "Entry level",      "Selenium"],
    ["Python Developer",      "Hyderabad", "Mid-Senior level", "Django"],
    ["DevOps Engineer",       "Pune",      "Associate",        "Kubernetes"],
    ["Data Engineer",         "Chennai",   "Mid-Senior level", "Apache Spark"],
]

for row_data in test_data:
    ws.append(row_data)

wb.save("src/test/resources/TestData.xlsx")
print("TestData.xlsx created at src/test/resources/TestData.xlsx")
